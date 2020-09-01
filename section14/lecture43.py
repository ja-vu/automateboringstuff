import openpyxl
import os

os.chdir('c:\\Workspace\\automateboringstuff\\section14')

wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'].value = 42
sheet['A2'].value = 'Hello'


sheet2 = wb.create_sheet()
sheet2.title = "My New Sheet"
wb.create_sheet(index=0, title="My other sheet")
print(wb.sheetnames)
wb.save('example.xlsx')
